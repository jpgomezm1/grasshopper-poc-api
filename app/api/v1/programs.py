"""Programs catalogue admin · GH-S8-BE-06/07/08.

Endpoints (super_admin only for mutations · public read for listing):

- POST   /programs                · create program
- GET    /programs                · list with pagination + filters
- GET    /programs/{id}           · detail
- PATCH  /programs/{id}           · partial update
- DELETE /programs/{id}           · soft (active=false) by default; hard with ?force=true
- POST   /programs/import         · upload Excel · validate + upsert (super_admin only)
- GET    /programs/export         · download canonical catalogue as Excel
"""
from __future__ import annotations

import io
import logging
import math
import unicodedata
import re
from datetime import datetime
from typing import Optional, List
from uuid import UUID

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    Request,
    Response,
    UploadFile,
    status,
)
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session as DBSession

from app.api.v1.auth import get_current_user
from app.config import get_settings
from app.core.rate_limiter import limiter
from app.db.database import get_db


def _rate_limit_programs_import(request: Request):
    """GH-S11-INFRA-04 · per-IP/user rate limit for program import upload."""
    from app.core.rate_limiter import rate_limit
    s = get_settings()
    return rate_limit(s.rate_limit_programs_import)(request)
from app.db.models import Program, User, UserRole
from app.schemas.program import (
    ProgramCreate,
    ProgramImportReport,
    ProgramListResponse,
    ProgramResponse,
    ProgramUpdate,
)
from app.schemas.roi import RoiCalculation
from app.services import roi_service
from app.services.audit_service import log_action

logger = logging.getLogger(__name__)


router = APIRouter(prefix="/programs", tags=["Programs"])


REQUIRED_FIELDS = [
    "program_id", "name", "slug", "country", "city", "institution",
    "type", "area", "subject", "duration_months", "cost_total",
    "currency", "budget_tier", "alliance_type", "active",
]


def _ensure_super_admin(user: User) -> None:
    if user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Forbidden · only super_admin can manage the program catalogue.",
        )


def _slugify(value: str) -> str:
    norm = unicodedata.normalize("NFKD", value or "")
    norm = norm.encode("ascii", "ignore").decode("ascii")
    norm = re.sub(r"[^a-zA-Z0-9]+", "-", norm).strip("-").lower()
    return norm or "program"


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _coerce_excel_float_optional(v):
    """float si la celda tiene valor numérico · None si vacía · raise si inválida.

    Permite columnas numéricas OPCIONALES en el import: vacío = "no tocar".
    """
    if _is_blank(v):
        return None
    if isinstance(v, str):
        v = v.strip().replace("%", "").replace(",", ".")
    return float(v)


def _coerce_excel_int_optional(v):
    """int si la celda tiene valor · None si vacía · raise si inválida."""
    if _is_blank(v):
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", "")
    return int(float(v))


_CEFR_LEVELS = {"A1", "A2", "B1", "B2", "C1", "C2"}


# (helper de string genérico no necesario · CEFR tiene su propio coercer)


def _coerce_excel_cefr_optional(v):
    """Nivel CEFR (A1..C2) en mayúsculas si la celda tiene valor · None si vacía.

    Lanza ValueError si el valor no es un nivel CEFR válido (lo marca como error
    de fila, no lo guarda silenciosamente mal).
    """
    if _is_blank(v):
        return None
    s = str(v).strip().upper()
    if s not in _CEFR_LEVELS:
        raise ValueError("nivel CEFR inválido (A1..C2)")
    return s


# D-002 · columnas OPCIONALES de admisión (Reach/Match/Safety). Cada una con su
# coerción. Vacío = no tocar (preserva lo curado). El valor se aplica al modelo
# solo si la columna venía en el Excel y la celda tenía contenido.
_ADMISSION_IMPORT_FIELDS = {
    "acceptance_rate": _coerce_excel_float_optional,
    "avg_admitted_gpa": _coerce_excel_float_optional,
    "min_sat": _coerce_excel_int_optional,
    "avg_sat": _coerce_excel_int_optional,
    "min_english_level": _coerce_excel_cefr_optional,
}


_TRUE_TOKENS = {"si", "sí", "yes", "true", "1", "y", "x"}


def _coerce_excel_bool(v, default: bool = False) -> bool:
    """Coerciona una celda de Excel a booleano (acepta si/sí/yes/true/1/y/x)."""
    if isinstance(v, str):
        return v.strip().lower() in _TRUE_TOKENS
    if v is None:
        return default
    return bool(v)


def _coerce_excel_bool_optional(v):
    """Como `_coerce_excel_bool` pero devuelve None si la celda está vacía.

    Permite columnas booleanas OPCIONALES en el import: una celda vacía = "no
    tocar este campo" (preserva el tri-estado True/False/desconocido), en vez de
    sobrescribir a False un valor curado previamente.
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().lower()
        if not s:
            return None
        return s in _TRUE_TOKENS
    return bool(v)


# ----------------------------- list / create / detail -----------------------------

@router.get(
    "",
    response_model=ProgramListResponse,
    summary="GH-S8-BE-06 · paginated list with filters",
)
def list_programs(
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=200),
    country: Optional[str] = Query(None),
    budget_tier: Optional[str] = Query(None),
    alliance_type: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    scholarships_for_latam: Optional[bool] = Query(
        None,
        description="F-003 · filtrar programas con/sin beca LatAm curada. None → no filtra.",
    ),
):
    q = db.query(Program)

    # school_admin / psychologist / student see only active programs
    if current_user.role != UserRole.SUPER_ADMIN:
        q = q.filter(Program.active.is_(True))
    elif active is not None:
        q = q.filter(Program.active.is_(active))

    if country:
        q = q.filter(Program.country == country)
    if budget_tier:
        q = q.filter(Program.budget_tier == budget_tier)
    if alliance_type:
        q = q.filter(Program.alliance_type == alliance_type)
    if type:
        q = q.filter(Program.type == type)
    if scholarships_for_latam is not None:
        q = q.filter(Program.scholarships_for_latam.is_(scholarships_for_latam))
    if search:
        term = f"%{search.strip().lower()}%"
        q = q.filter(
            or_(
                Program.name.ilike(term),
                Program.institution.ilike(term),
                Program.subject.ilike(term),
                Program.program_id.ilike(term),
            )
        )

    total = q.count()
    rows = (
        q.order_by(Program.country.asc(), Program.institution.asc(), Program.name.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    total_pages = max(1, math.ceil(total / page_size)) if total else 0

    return ProgramListResponse(
        items=[ProgramResponse.model_validate(r) for r in rows],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
    )


@router.post(
    "",
    response_model=ProgramResponse,
    status_code=status.HTTP_201_CREATED,
    summary="GH-S8-BE-06 · create program (super_admin only)",
)
def create_program(
    payload: ProgramCreate,
    request: Request,
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_super_admin(current_user)

    program = Program(
        program_id=payload.program_id.strip(),
        name=payload.name.strip(),
        slug=payload.slug,
        country=payload.country.strip(),
        city=(payload.city or "").strip() or None,
        institution=payload.institution.strip(),
        type=payload.type.strip(),
        area=(payload.area or "").strip() or None,
        subject=(payload.subject or "").strip() or None,
        duration_months=payload.duration_months,
        cost_total=payload.cost_total,
        currency=payload.currency,
        budget_tier=payload.budget_tier,
        alliance_type=payload.alliance_type,
        language_requirement=payload.language_requirement,
        active=payload.active,
        # D-002 · variables de admisión (None si no se curan)
        acceptance_rate=payload.acceptance_rate,
        avg_admitted_gpa=payload.avg_admitted_gpa,
        min_sat=payload.min_sat,
        avg_sat=payload.avg_sat,
        min_english_level=payload.min_english_level,
        # F-003 · beca curada para LatAm
        scholarships_for_latam=payload.scholarships_for_latam,
        raw=payload.raw,
    )
    db.add(program)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="A program with the same program_id or slug already exists.",
        )
    db.refresh(program)

    log_action(
        db,
        user=current_user,
        action="program.create",
        resource_type="program",
        resource_id=str(program.id),
        payload={"program_id": program.program_id, "slug": program.slug},
        request=request,
    )

    return ProgramResponse.model_validate(program)


@router.get(
    "/export.xlsx",
    summary="GH-S8-BE-08 · download canonical program catalogue as Excel",
)
def export_programs(
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_super_admin(current_user)

    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl missing on server.")

    rows = db.query(Program).order_by(Program.country.asc(), Program.institution.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Programs"
    ws.append(
        REQUIRED_FIELDS
        + ["language_requirement", "scholarships_for_latam"]
        + list(_ADMISSION_IMPORT_FIELDS.keys())
    )
    for p in rows:
        # F-003 · tri-estado: vacío = desconocido (sin curar) · si/no = curado
        if p.scholarships_for_latam is None:
            beca = ""
        else:
            beca = "si" if p.scholarships_for_latam else "no"
        ws.append([
            p.program_id,
            p.name,
            p.slug,
            p.country,
            p.city or "",
            p.institution,
            p.type,
            p.area or "",
            p.subject or "",
            p.duration_months,
            p.cost_total,
            p.currency,
            p.budget_tier,
            p.alliance_type,
            "si" if p.active else "no",
            p.language_requirement or "",
            # F-003 · tri-estado: vacío = sin curar
            beca,
            # D-002 · vacío = sin curar (preserva tri-estado en el round-trip)
            p.acceptance_rate if p.acceptance_rate is not None else "",
            p.avg_admitted_gpa if p.avg_admitted_gpa is not None else "",
            p.min_sat if p.min_sat is not None else "",
            p.avg_sat if p.avg_sat is not None else "",
            p.min_english_level or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"grasshopper_catalog_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/{program_id}",
    response_model=ProgramResponse,
    summary="Read program by id",
)
def get_program(
    program_id: UUID,
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    program = db.query(Program).filter(Program.id == program_id).first()
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Program not found.")
    if not program.active and current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Program not found.")
    return ProgramResponse.model_validate(program)


@router.get(
    "/{program_id}/roi",
    response_model=RoiCalculation,
    summary="F-002 · ROI calculator (visa + cost + post-grad earnings)",
)
def get_program_roi(
    program_id: UUID,
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return the ROI calculation for a program.

    GH-LOCAL-CLIENT-MODULES · F-002 etapa 1 · 2026-05-21.

    Combines: tuition cost · living cost in destination city · entry salary ·
    visa work years. Returns payback in years + net value + rating.

    Authenticated users only. Returns 404 if program doesn't exist or is
    inactive (and current_user is not super_admin).
    """
    program = db.query(Program).filter(Program.id == program_id).first()
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Program not found.")
    if not program.active and current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Program not found.")
    return roi_service.calculate_roi(program)


@router.patch(
    "/{program_id}",
    response_model=ProgramResponse,
    summary="GH-S8-BE-06 · update program (super_admin only)",
)
def update_program(
    program_id: UUID,
    payload: ProgramUpdate,
    request: Request,
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_super_admin(current_user)

    program = db.query(Program).filter(Program.id == program_id).first()
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Program not found.")

    diff: dict = {}
    for fld, val in payload.model_dump(exclude_unset=True).items():
        cur = getattr(program, fld)
        if val != cur:
            diff[fld] = {"from": str(cur)[:80] if cur is not None else None, "to": str(val)[:80] if val is not None else None}
            setattr(program, fld, val)

    program.updated_at = datetime.utcnow()
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Slug or program_id collision.",
        )
    db.refresh(program)

    if diff:
        log_action(
            db,
            user=current_user,
            action="program.update",
            resource_type="program",
            resource_id=str(program.id),
            payload=diff,
            request=request,
        )

    return ProgramResponse.model_validate(program)


@router.delete(
    "/{program_id}",
    status_code=status.HTTP_200_OK,
    summary="GH-S8-BE-06 · soft delete program (active=false) · ?force=true for hard delete",
)
def delete_program(
    program_id: UUID,
    request: Request,
    force: bool = Query(False),
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_super_admin(current_user)

    program = db.query(Program).filter(Program.id == program_id).first()
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Program not found.")

    if force:
        db.delete(program)
        action = "program.delete"
    else:
        program.active = False
        program.updated_at = datetime.utcnow()
        action = "program.delete"
    db.commit()

    log_action(
        db,
        user=current_user,
        action=action,
        resource_type="program",
        resource_id=str(program_id),
        payload={"force": force, "program_biz_id": program.program_id},
        request=request,
    )

    return {"id": str(program_id), "deleted": force, "deactivated": not force}


# ----------------------------- import / export -----------------------------

@router.post(
    "/import",
    response_model=ProgramImportReport,
    summary="GH-S8-BE-07 · Excel upload · validate + upsert (super_admin only)",
    dependencies=[Depends(_rate_limit_programs_import)],
)
async def import_programs(
    request: Request,
    file: UploadFile = File(...),
    commit: bool = Query(False, description="false = dry-run (validate only)"),
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """GH-S11-INFRA-04 · rate-limited (default 5/hour)."""
    _ensure_super_admin(current_user)

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl missing on server.",
        )

    if file.content_type and "spreadsheetml" not in file.content_type and not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are accepted.",
        )

    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:  # 10 MB
        raise HTTPException(status_code=413, detail="Excel too large (>10MB).")

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty Excel.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = [c for c in REQUIRED_FIELDS if c not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {missing}")

    errors: list = []
    warnings: list = []
    valid_records: list[dict] = []
    total_rows = 0
    for idx, raw_row in enumerate(rows[1:], start=2):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw_row):
            continue
        total_rows += 1
        record = dict(zip(headers, raw_row))
        # normalize a few fields
        for k in ("program_id", "name", "country", "city", "institution", "type", "area", "subject", "currency", "budget_tier", "alliance_type", "language_requirement", "slug"):
            if record.get(k) is not None:
                record[k] = str(record[k]).strip()
        # required presence
        row_ok = True
        for fld in REQUIRED_FIELDS:
            v = record.get(fld)
            if v is None or (isinstance(v, str) and not v.strip()):
                errors.append({"row": idx, "field": fld, "msg": "campo obligatorio vacío"})
                row_ok = False
        # numeric coercions
        try:
            record["duration_months"] = int(record.get("duration_months") or 0)
        except (TypeError, ValueError):
            errors.append({"row": idx, "field": "duration_months", "msg": "no es entero"})
            row_ok = False
        try:
            record["cost_total"] = int(record.get("cost_total") or 0)
        except (TypeError, ValueError):
            errors.append({"row": idx, "field": "cost_total", "msg": "no es entero"})
            row_ok = False
        # active boolean coercion (default True si la celda viene vacía)
        record["active"] = _coerce_excel_bool(record.get("active"), default=True)

        # F-003 · columna OPCIONAL de beca LatAm · vacío = no tocar
        if "scholarships_for_latam" in headers:
            record["scholarships_for_latam"] = _coerce_excel_bool_optional(
                record.get("scholarships_for_latam")
            )

        # D-002 · columnas OPCIONALES de admisión · vacío = no tocar
        for col, coerce in _ADMISSION_IMPORT_FIELDS.items():
            if col in headers:
                try:
                    record[col] = coerce(record.get(col))
                except (TypeError, ValueError):
                    errors.append({"row": idx, "field": col, "msg": "valor inválido"})
                    row_ok = False

        if row_ok:
            valid_records.append(record)

    inserted = 0
    updated = 0
    if commit and not errors:
        for r in valid_records:
            existing = db.query(Program).filter(Program.program_id == r["program_id"]).first()
            slug = r.get("slug") or _slugify(r["name"])
            if existing:
                # update fields
                existing.name = r["name"]
                existing.slug = slug
                existing.country = r["country"]
                existing.city = r.get("city") or None
                existing.institution = r["institution"]
                existing.type = r["type"]
                existing.area = r.get("area") or None
                existing.subject = r.get("subject") or None
                existing.duration_months = r["duration_months"]
                existing.cost_total = r["cost_total"]
                existing.currency = (r.get("currency") or "USD").upper()
                existing.budget_tier = (r.get("budget_tier") or "medium").lower()
                existing.alliance_type = (r.get("alliance_type") or "estandar").lower()
                existing.language_requirement = r.get("language_requirement") or None
                existing.active = bool(r.get("active", True))
                # F-003 · solo si la columna venía y la celda tenía valor
                if "scholarships_for_latam" in headers and r.get("scholarships_for_latam") is not None:
                    existing.scholarships_for_latam = r["scholarships_for_latam"]
                # D-002 · solo si la columna venía y la celda tenía valor
                for col in _ADMISSION_IMPORT_FIELDS:
                    if col in headers and r.get(col) is not None:
                        setattr(existing, col, r[col])
                existing.raw = r
                existing.updated_at = datetime.utcnow()
                updated += 1
            else:
                p = Program(
                    program_id=r["program_id"],
                    name=r["name"],
                    slug=slug,
                    country=r["country"],
                    city=r.get("city") or None,
                    institution=r["institution"],
                    type=r["type"],
                    area=r.get("area") or None,
                    subject=r.get("subject") or None,
                    duration_months=r["duration_months"],
                    cost_total=r["cost_total"],
                    currency=(r.get("currency") or "USD").upper(),
                    budget_tier=(r.get("budget_tier") or "medium").lower(),
                    alliance_type=(r.get("alliance_type") or "estandar").lower(),
                    language_requirement=r.get("language_requirement") or None,
                    active=bool(r.get("active", True)),
                    scholarships_for_latam=(
                        r.get("scholarships_for_latam")
                        if "scholarships_for_latam" in headers
                        else None
                    ),
                    raw=r,
                )
                # D-002 · campos de admisión opcionales (si la columna venía)
                for col in _ADMISSION_IMPORT_FIELDS:
                    if col in headers and r.get(col) is not None:
                        setattr(p, col, r[col])
                db.add(p)
                inserted += 1
        try:
            db.commit()
        except IntegrityError as exc:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Integrity error: {exc.orig}")

        log_action(
            db,
            user=current_user,
            action="program.import",
            resource_type="program_catalog",
            resource_id=None,
            payload={
                "filename": file.filename,
                "inserted": inserted,
                "updated": updated,
                "total_rows": total_rows,
            },
            request=request,
        )

    return ProgramImportReport(
        total_rows=total_rows,
        valid_rows=len(valid_records),
        inserted=inserted,
        updated=updated,
        errors=errors,
        warnings=warnings,
        committed=bool(commit and not errors),
    )


# ---------------------------------------------------------------------------
# Bloque B · upload helpers (Sprint super_admin fixes 2026-05-03)
# ---------------------------------------------------------------------------

ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}
ALLOWED_IMAGE_EXTS = {"jpg", "jpeg", "png", "webp"}
MAX_IMAGE_MB = 5


def _safe_image_upload(
    upload: UploadFile, *, kind: str, program_biz_id: str
) -> tuple[bytes, str, str]:
    """Validate the upload + return (data, content_type, filename).

    `kind` is "images" or "logos" → used to scope the storage path.
    Defends against:
      - oversized payloads (5MB cap)
      - unknown content types (only JPG/PNG/WebP)
      - path traversal in filename (we sanitize via storage_service)
    """
    if upload.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tipo no permitido · usa JPG/PNG/WebP. Recibido: {upload.content_type}",
        )

    data = upload.file.read()
    if len(data) > MAX_IMAGE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Archivo excede {MAX_IMAGE_MB}MB.",
        )

    # filename: keep extension only, prefix with timestamp + biz id
    raw_name = (upload.filename or "image").lower()
    ext = raw_name.rsplit(".", 1)[-1] if "." in raw_name else ""
    if ext not in ALLOWED_IMAGE_EXTS:
        # try to infer from content-type
        ext = {
            "image/jpeg": "jpg",
            "image/png": "png",
            "image/webp": "webp",
        }[upload.content_type]
    safe_filename = (
        f"{program_biz_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{kind}.{ext}"
    )
    return data, upload.content_type, safe_filename


@router.post(
    "/{program_id}/upload-image",
    summary="Bloque B · upload editorial image (super_admin · max 5MB JPG/PNG/WebP)",
)
def upload_program_image(
    program_id: UUID,
    request: Request,
    file: UploadFile = File(...),
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload one image for the program gallery.

    Stores the bytes via `storage_service` (stub or supabase depending on env)
    and returns the path + signed URL so the FE can render. Does NOT mutate
    `programs.images` automatically · the FE composes the JSON list and PATCHes
    the row to keep the order and metadata under explicit user control.
    """
    _ensure_super_admin(current_user)
    program = db.query(Program).filter(Program.id == program_id).first()
    if not program:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Program not found."
        )

    from app.services import storage_service

    data, content_type, filename = _safe_image_upload(
        file, kind="image", program_biz_id=program.program_id
    )
    path = storage_service.build_user_path(
        f"program_{program.id}", "images", filename
    )
    storage_service.upload_file(
        path=path, data=data, content_type=content_type, max_size_mb=MAX_IMAGE_MB
    )
    signed = storage_service.get_signed_url(path, expires_in_seconds=60 * 60 * 24)

    log_action(
        db,
        user=current_user,
        action="program.upload_image",
        resource_type="program",
        resource_id=str(program.id),
        payload={"filename": filename, "size": len(data)},
        request=request,
    )

    return {
        "path": path,
        "url": signed,
        "filename": filename,
        "content_type": content_type,
        "size_bytes": len(data),
    }


@router.post(
    "/{program_id}/upload-logo",
    summary="Bloque B · upload institution logo (super_admin · max 5MB)",
)
def upload_program_logo(
    program_id: UUID,
    request: Request,
    file: UploadFile = File(...),
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload the institution logo and update programs.institution_logo_url.

    Side effect: persists the signed URL on the row. Note the URL is signed
    (1 day TTL) · for permanent display the FE can re-fetch via /programs/:id
    or rely on Supabase public bucket configuration in production.
    """
    _ensure_super_admin(current_user)
    program = db.query(Program).filter(Program.id == program_id).first()
    if not program:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Program not found."
        )

    from app.services import storage_service

    data, content_type, filename = _safe_image_upload(
        file, kind="logo", program_biz_id=program.program_id
    )
    path = storage_service.build_user_path(
        f"program_{program.id}", "logos", filename
    )
    storage_service.upload_file(
        path=path, data=data, content_type=content_type, max_size_mb=MAX_IMAGE_MB
    )
    signed = storage_service.get_signed_url(path, expires_in_seconds=60 * 60 * 24)

    program.institution_logo_url = signed
    program.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(program)

    log_action(
        db,
        user=current_user,
        action="program.upload_logo",
        resource_type="program",
        resource_id=str(program.id),
        payload={"filename": filename, "size": len(data)},
        request=request,
    )

    return {
        "path": path,
        "url": signed,
        "institution_logo_url": program.institution_logo_url,
    }


@router.get(
    "/by-slug/{slug}",
    response_model=ProgramResponse,
    summary="Bloque B · public program detail by slug (any authenticated user)",
)
def get_program_by_slug(
    slug: str,
    db: DBSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Public-facing detail · used by the editorial detail page (B2C/B2B).

    Returns 404 for inactive programs unless the caller is super_admin.
    """
    program = (
        db.query(Program).filter(Program.slug == slug.lower().strip()).first()
    )
    if not program:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Program not found."
        )
    if not program.active and current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Program not found."
        )
    return ProgramResponse.model_validate(program)
