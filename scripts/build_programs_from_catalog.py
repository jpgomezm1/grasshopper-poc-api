"""GH-CATALOG-REAL · Construye `programs` (ofertas del alumno) desde el
catálogo REAL de instituciones del xlsx del cliente (2026-06-03).

Motivación: hasta hoy la tabla `programs` (lo que ve el alumno en /ofertas)
era seed inventado (precios/duraciones falsos). El cliente confirmó que la
fuente real de instituciones es el xlsx (mismo que alimenta
`institutions_catalog`), pero ese xlsx NO trae precio, duración ni ROI de cara
al estudiante — sólo institución, país/ciudad, categoría, programas ofrecidos
y datos de contrato/comisión (estos últimos SENSIBLES, no se exponen).

Por eso este importador:
    - Genera 1 `Program` por institución real (grano institución).
    - Llena SÓLO datos reales: nombre, país, ciudad, institución, tipo
      (derivado de categoría/programas), website, tags.
    - Deja cost_total / duration_months / budget_tier en NULL = "a confirmar"
      (requiere migration 048). El ROI queda oculto hasta tener cifras.
    - NUNCA copia comisiones ni contactos al lado del alumno.

Uso:
    python scripts/build_programs_from_catalog.py <ruta-al-xlsx>            # dry-run
    python scripts/build_programs_from_catalog.py <ruta-al-xlsx> --commit   # escribe DB
    python scripts/build_programs_from_catalog.py <ruta-al-xlsx> --commit --deactivate-seed
        # además marca active=False los programs que NO vienen del catálogo (seed viejo)

Idempotente: upsert por `program_id` (slug de la institución).
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl no instalado. Ejecutar: pip install openpyxl")

# Reusamos el parser robusto del importador de instituciones (mismo xlsx).
from scripts.import_institutions import (  # noqa: E402
    parse_instituciones,
    parse_resumen,
    dedupe_and_merge,
    ImportReport,
    _str,
    _norm_country,
    _norm_category,
    _parse_programs,
)

# Categoría del catálogo → Program.type (valores que entiende el mapeo
# Program→Oferta en app/api/v1/ofertas.py · _TYPE_TO_CATEGORY).
_CATEGORY_TO_TYPE = {
    "Universidad": "pregrado",
    "Business School": "maestria",
    "Instituto Idiomas": "vacacional",       # → curso_idiomas
    "College Privado": "curso_corto",        # → certificacion_corta
    "College Público": "curso_corto",
    "Polytechnic": "curso_corto",
    "High School": "intercambio",            # → semestre_academico
    "Summer School": "vacacional",
    "Camps": "vacacional",
    "Proveedor": "curso_corto",
}
# Pistas por programa ofrecido (texto libre) cuando la categoría no alcanza.
_PROGRAM_HINTS = (
    ("idioma", "vacacional"),
    ("foundation", "pregrado"),
    ("pre master", "maestria"),
    ("pregrado", "pregrado"),
    ("undergrad", "pregrado"),
    ("master", "maestria"),
    ("posgrado", "maestria"),
    ("mba", "mba"),
    ("doctor", "doctorado"),
    ("vocacion", "curso_corto"),
    ("diplom", "diplomado"),
    ("intercambio", "intercambio"),
    ("year", "intercambio"),
)


def _slugify(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s or "programa"


def _derive_type(category: Optional[str], programs: List[str]) -> str:
    if category and category in _CATEGORY_TO_TYPE:
        base = _CATEGORY_TO_TYPE[category]
    else:
        base = None
    joined = " ".join(programs).lower()
    for needle, t in _PROGRAM_HINTS:
        if needle in joined:
            return t
    return base or "curso_corto"


def _build_program(rec: Dict[str, Any], used_ids: set) -> Optional[Dict[str, Any]]:
    name = rec.get("name")
    country = rec.get("country")
    if not name or not country:
        return None  # sin país no podemos ubicar la oferta
    programs = rec.get("programs_offered") or []
    ptype = _derive_type(rec.get("category"), programs)

    pid = _slugify(name)
    base_pid = pid
    i = 2
    while pid in used_ids:
        pid = f"{base_pid}-{i}"
        i += 1
    used_ids.add(pid)

    progs_txt = ", ".join(programs) if programs else "Consultar con tu advisor"
    desc = (
        f"{name} es una institución del catálogo oficial de Grasshopper en "
        f"{rec.get('city') + ', ' if rec.get('city') else ''}{country}. "
        f"Programas disponibles: {progs_txt}. "
        f"El precio y la duración exactos se confirman con tu advisor según el "
        f"programa y la fecha de inicio."
    )
    tags = [t for t in [rec.get("category"), country, rec.get("partner_group")] if t]

    return {
        "program_id": pid,
        "name": name,
        "slug": pid,
        "country": country,
        "city": rec.get("city"),
        "institution": name,
        "type": ptype,
        "area": None,
        "subject": (programs[0] if programs else None),
        # --- financieros REALES desconocidos → NULL ("a confirmar") ---
        "duration_months": None,
        "cost_total": None,
        "currency": "USD",
        "budget_tier": None,
        "alliance_type": "estandar",
        "language_requirement": None,
        "active": bool(rec.get("active", True)),
        "description_long": desc,
        "tags": tags,
        "highlights": [f"Programas: {progs_txt}"] if programs else [],
        # trazabilidad (NO comisiones / NO contactos)
        "raw": {
            "source": "institutions_catalog",
            "category": rec.get("category"),
            "website": rec.get("website"),
            "partner_group": rec.get("partner_group"),
            "programs_offered": programs,
        },
    }


def _rec(
    name: str,
    country: Optional[str],
    *,
    city: Optional[str] = None,
    category: Optional[str] = None,
    programs: Optional[List[str]] = None,
    website: Optional[str] = None,
    group: Optional[str] = None,
    source: str = "",
) -> Dict[str, Any]:
    return {
        "name": name,
        "category": category,
        "country": country,
        "city": city,
        "partner_group": group,
        "programs_offered": programs or [],
        "agreement_status": None,
        "website": website,
        "source_sheet": source,
        "active": True,
    }


def _parse_provider(ws, source: str, cols: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Parser genérico para hojas de aliados.

    ``cols`` mapea campo→índice de columna; ``country_default`` fija el país
    cuando la hoja no trae columna de país (p.ej. EdAgent=Australia).
    Las columnas de comisión/contacto se IGNORAN deliberadamente.
    """
    out: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or all(v is None for v in r):
            continue
        ni = cols["name"]
        name = _str(r[ni]) if len(r) > ni else None
        if not name or name.startswith("#"):
            continue
        ci = cols.get("country")
        country = None
        if ci is not None and len(r) > ci:
            country = _norm_country(r[ci])
        country = country or cols.get("country_default")
        cat = None
        if cols.get("category") is not None and len(r) > cols["category"]:
            cat = _norm_category(r[cols["category"]])
        city = None
        if cols.get("city") is not None and len(r) > cols["city"]:
            city = _str(r[cols["city"]])
        progs: List[str] = []
        for pi in cols.get("programs", []):
            if len(r) > pi:
                progs += _parse_programs(r[pi])
        web = None
        if cols.get("website") is not None and len(r) > cols["website"]:
            web = _str(r[cols["website"]])
        grp = None
        if cols.get("group") is not None and len(r) > cols["group"]:
            grp = _str(r[cols["group"]])
        out.append(_rec(name, country, city=city, category=cat, programs=progs, website=web, group=grp, source=source))
    return out


# Layout de columnas por hoja de aliado (0-indexed). Sólo campos útiles al alumno.
_PROVIDER_LAYOUTS = {
    "INTO":                 {"name": 0, "group": 2, "country": 3, "city": 4, "programs": [9, 10, 11, 12, 13]},
    "Study Group":          {"name": 0, "group": 2, "country": 3, "city": 4, "programs": [9, 10, 11, 12, 13]},
    "Shorelight":           {"name": 0, "group": 2, "country": 3, "city": 4, "programs": [5, 6, 7, 8, 9]},
    "Oxford International":  {"name": 0, "group": 2, "country": 3, "city": 4, "programs": [5, 6, 7, 8, 9]},
    "Kaplan HED":           {"name": 0, "group": 2, "country": 3, "city": 4, "programs": [5, 6, 7, 8, 9]},
    "Wellsprings":          {"name": 0, "group": 2, "country": 3, "city": 4, "programs": [5, 6, 7, 8, 9]},
    "Applyboard":           {"name": 3, "country": 1, "city": 0, "category": 2},  # City,Country,EstablishmentType,Name,Province,SubmissionType
    "colleges Canada":      {"name": 0, "country_default": "Canada"},             # ignora ESTADO + comisiones
    "EdAgent AMET- Australia": {"name": 0, "website": 1, "country_default": "Australia"},
    "Hoja1 (2)":            {"name": 1, "country": 2, "category": 0},             # Type,Name,Country
}


def build(xlsx_path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    sources: List[Any] = []
    # 1) Convenios directos (fuente de verdad, gana en dedupe) · excluye "Vencido".
    if "Instituciones" in wb.sheetnames:
        direct = parse_instituciones(wb["Instituciones"], ImportReport(sheet="Instituciones"))
        direct = [r for r in direct if (r.get("agreement_status") or "").strip().lower() != "vencido"]
        sources.append(("Instituciones", direct))
    if "Instituciones Resumen" in wb.sheetnames:
        sources.append(("Instituciones Resumen", parse_resumen(wb["Instituciones Resumen"], ImportReport(sheet="Instituciones Resumen"))))
    # 2) Hojas de aliados (acceso vía agregador) · entran completas, dedup por nombre.
    for sheet, layout in _PROVIDER_LAYOUTS.items():
        if sheet in wb.sheetnames:
            sources.append((sheet, _parse_provider(wb[sheet], sheet, layout)))

    merged, _ = dedupe_and_merge(sources)
    used: set = set()
    out: List[Dict[str, Any]] = []
    for rec in merged:
        prog = _build_program(rec, used)
        if prog:
            out.append(prog)
    return out


def commit_to_db(records: List[Dict[str, Any]], deactivate_seed: bool) -> Dict[str, int]:
    from app.db.database import SessionLocal
    from app.db.models import Program

    stats = {"inserted": 0, "updated": 0, "deactivated_seed": 0}
    catalog_ids = {r["program_id"] for r in records}
    with SessionLocal() as session:
        existing = {p.program_id: p for p in session.query(Program).all()}
        for rec in records:
            obj = existing.get(rec["program_id"])
            if obj is None:
                session.add(Program(**rec))
                stats["inserted"] += 1
            else:
                for k, v in rec.items():
                    setattr(obj, k, v)
                stats["updated"] += 1
        if deactivate_seed:
            for pid, obj in existing.items():
                if pid not in catalog_ids and obj.active:
                    obj.active = False
                    stats["deactivated_seed"] += 1
        session.commit()
    return stats


def main() -> int:
    ap = argparse.ArgumentParser(description="Construye programs desde el catálogo real (xlsx).")
    ap.add_argument("xlsx_path", type=Path)
    ap.add_argument("--commit", action="store_true", help="Escribir a DB (default dry-run).")
    ap.add_argument("--deactivate-seed", action="store_true", help="Marca active=False los programs que no vienen del catálogo.")
    args = ap.parse_args()

    if not args.xlsx_path.exists():
        print(f"[ERROR] No existe: {args.xlsx_path}", file=sys.stderr)
        return 2

    print(f"[INFO] Construyendo programs desde {args.xlsx_path.name}...")
    records = build(args.xlsx_path)
    print(f"[INFO] {len(records)} programas reales construidos (1 por institución con país).")

    # Breakdown por tipo y país
    by_type: Dict[str, int] = {}
    by_country: Dict[str, int] = {}
    for r in records:
        by_type[r["type"]] = by_type.get(r["type"], 0) + 1
        by_country[r["country"]] = by_country.get(r["country"], 0) + 1
    print("\nPor tipo:", dict(sorted(by_type.items(), key=lambda x: -x[1])))
    print("Top países:", dict(sorted(by_country.items(), key=lambda x: -x[1])[:10]))
    print("\nMuestra (primeros 5):")
    for r in records[:5]:
        print(f"  · {r['name']} [{r['type']}] · {r['city'] or '-'}, {r['country']} · cost/dur=NULL")

    if not args.commit:
        print("\n[OK] Dry-run · NO se escribió a DB. Re-correr con --commit para persistir.")
        return 0

    print(f"\n[INFO] Escribiendo {len(records)} programs a DB (deactivate_seed={args.deactivate_seed})...")
    stats = commit_to_db(records, args.deactivate_seed)
    print(f"[OK] {stats}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
