"""
GH-S1-BE-02 · Import de catalogo de programas (Excel -> tabla `programs`).
GH-S1-BE-03 · Validacion: reporta filas con datos incompletos antes de cargar.

Uso:
    cd backend
    python scripts/import_catalog.py samples/catalog_dummy.xlsx
    python scripts/import_catalog.py path/al/catalogo_real.xlsx --commit

Por defecto corre en modo --dry-run (NO escribe en DB · solo valida).
Pasar --commit para confirmar la importacion.

Nota: La tabla `programs` se crea formalmente en GH-S8-BE-06. Mientras tanto
este script valida exclusivamente. El upsert SQL queda preparado y se activa
cuando exista la tabla.
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl no instalado. Ejecutar: pip install openpyxl")


REQUIRED_FIELDS = {
    "program_id", "name", "slug", "country", "city", "institution",
    "type", "area", "subject", "duration_months", "cost_total",
    "currency", "budget_tier", "alliance_type", "active",
}

VALID_TIERS = {"low", "medium", "high", "premium"}
VALID_ALLIANCES = {"preferencial", "estandar", "convenio"}
VALID_CURRENCIES = {"USD", "EUR", "GBP", "CAD", "AUD", "CHF"}


@dataclass
class ValidationReport:
    total_rows: int = 0
    valid_rows: int = 0
    errors: list[tuple[int, str, str]] = field(default_factory=list)  # (row_number, field, msg)
    warnings: list[tuple[int, str, str]] = field(default_factory=list)

    def add_error(self, row: int, field_name: str, message: str) -> None:
        self.errors.append((row, field_name, message))

    def add_warning(self, row: int, field_name: str, message: str) -> None:
        self.warnings.append((row, field_name, message))

    def is_valid(self) -> bool:
        return len(self.errors) == 0

    def print(self) -> None:
        print(f"\n=== Reporte de validacion ===")
        print(f"Filas totales:  {self.total_rows}")
        print(f"Filas validas:  {self.valid_rows}")
        print(f"Errores:        {len(self.errors)}")
        print(f"Warnings:       {len(self.warnings)}")
        if self.errors:
            print("\n--- ERRORES (bloqueantes) ---")
            for row, fld, msg in self.errors[:50]:
                print(f"  fila {row} · campo `{fld}` · {msg}")
            if len(self.errors) > 50:
                print(f"  ...y {len(self.errors) - 50} mas")
        if self.warnings:
            print("\n--- WARNINGS (no bloqueantes) ---")
            for row, fld, msg in self.warnings[:20]:
                print(f"  fila {row} · campo `{fld}` · {msg}")
            if len(self.warnings) > 20:
                print(f"  ...y {len(self.warnings) - 20} mas")


def validate_row(row_idx: int, data: dict, report: ValidationReport) -> bool:
    ok = True
    # required fields presence
    for fld in REQUIRED_FIELDS:
        val = data.get(fld)
        if val is None or (isinstance(val, str) and not val.strip()):
            report.add_error(row_idx, fld, "campo obligatorio vacio")
            ok = False

    # enum validations
    tier = data.get("budget_tier")
    if tier and str(tier).strip() not in VALID_TIERS:
        report.add_error(row_idx, "budget_tier", f"valor invalido `{tier}` · esperado uno de {sorted(VALID_TIERS)}")
        ok = False

    alliance = data.get("alliance_type")
    if alliance and str(alliance).strip() not in VALID_ALLIANCES:
        report.add_error(row_idx, "alliance_type", f"valor invalido `{alliance}` · esperado uno de {sorted(VALID_ALLIANCES)}")
        ok = False

    currency = data.get("currency")
    if currency and str(currency).strip() not in VALID_CURRENCIES:
        report.add_warning(row_idx, "currency", f"moneda no estandar `{currency}` · revisar")

    # numeric checks
    dur = data.get("duration_months")
    if dur is not None:
        try:
            d = int(dur)
            if d <= 0 or d > 96:
                report.add_warning(row_idx, "duration_months", f"duracion sospechosa: {d} meses")
        except (TypeError, ValueError):
            report.add_error(row_idx, "duration_months", f"no es entero: `{dur}`")
            ok = False

    cost = data.get("cost_total")
    if cost is not None:
        try:
            c = int(cost)
            if c < 0:
                report.add_error(row_idx, "cost_total", "costo negativo")
                ok = False
        except (TypeError, ValueError):
            report.add_error(row_idx, "cost_total", f"no es entero: `{cost}`")
            ok = False

    return ok


def parse_excel(path: Path) -> tuple[list[dict], list[str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Hoja vacia en {path}")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data: list[dict] = []
    for raw in rows[1:]:
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        record = dict(zip(headers, raw))
        data.append(record)
    return data, headers


def upsert_programs(records: list[dict]) -> int:
    """Upsert masivo en tabla `programs`. La tabla se crea en GH-S8-BE-06.

    Mientras la tabla no exista, esta funcion solo loguea lo que insertaria.
    """
    try:
        # imports diferidos para que el script corra incluso si el backend no esta instalado
        from sqlalchemy import inspect, text  # noqa
        from app.db.database import SessionLocal, engine
    except Exception as exc:  # pragma: no cover
        print(f"[WARN] No se pudo importar SQLAlchemy/DB: {exc}")
        print(f"[WARN] Skipping insercion. {len(records)} registros NO insertados.")
        return 0

    inspector = inspect(engine)
    if "programs" not in inspector.get_table_names():
        print("[INFO] Tabla `programs` aun no existe (se crea en GH-S8-BE-06).")
        print(f"[INFO] {len(records)} registros validos · listos para insertar cuando exista la tabla.")
        return 0

    inserted = 0
    with SessionLocal() as session:
        # placeholder upsert · ajustar columnas exactas en S8
        for r in records:
            session.execute(
                text(
                    "INSERT INTO programs (program_id, name, slug, country, city, institution, "
                    "type, area, subject, duration_months, cost_total, currency, budget_tier, "
                    "alliance_type, active, raw) VALUES "
                    "(:program_id, :name, :slug, :country, :city, :institution, :type, :area, "
                    ":subject, :duration_months, :cost_total, :currency, :budget_tier, "
                    ":alliance_type, :active, :raw) "
                    "ON CONFLICT (program_id) DO UPDATE SET "
                    "name=excluded.name, slug=excluded.slug, country=excluded.country, "
                    "city=excluded.city, institution=excluded.institution, type=excluded.type, "
                    "area=excluded.area, subject=excluded.subject, "
                    "duration_months=excluded.duration_months, cost_total=excluded.cost_total, "
                    "currency=excluded.currency, budget_tier=excluded.budget_tier, "
                    "alliance_type=excluded.alliance_type, active=excluded.active, raw=excluded.raw"
                ),
                {**{k: r.get(k) for k in [
                    "program_id", "name", "slug", "country", "city", "institution",
                    "type", "area", "subject", "duration_months", "cost_total",
                    "currency", "budget_tier", "alliance_type", "active",
                ]}, "raw": str(r)},
            )
            inserted += 1
        session.commit()
    return inserted


def main() -> int:
    parser = argparse.ArgumentParser(description="Import + validate Grasshopper program catalog")
    parser.add_argument("excel_path", type=Path, help="Path al archivo Excel del catalogo")
    parser.add_argument("--commit", action="store_true", help="Insertar en DB (default es dry-run)")
    parser.add_argument("--strict", action="store_true", help="Si hay warnings, fallar")
    args = parser.parse_args()

    if not args.excel_path.exists():
        print(f"[ERROR] No existe el archivo: {args.excel_path}", file=sys.stderr)
        return 2

    print(f"[INFO] Cargando {args.excel_path}...")
    records, headers = parse_excel(args.excel_path)
    print(f"[INFO] {len(records)} filas leidas · {len(headers)} columnas")

    missing_cols = REQUIRED_FIELDS - set(headers)
    if missing_cols:
        print(f"[ERROR] Faltan columnas obligatorias: {sorted(missing_cols)}", file=sys.stderr)
        return 2

    report = ValidationReport(total_rows=len(records))
    valid_records: list[dict] = []
    for idx, rec in enumerate(records, start=2):  # fila 2 es la primera de datos
        if validate_row(idx, rec, report):
            report.valid_rows += 1
            valid_records.append(rec)

    report.print()

    if not report.is_valid():
        print("\n[FAIL] Validacion fallo · NO se insertan registros.")
        return 1
    if args.strict and report.warnings:
        print("\n[FAIL] Strict mode · warnings presentes · NO se insertan registros.")
        return 1

    if args.commit:
        print(f"\n[INFO] Insertando {len(valid_records)} programas en DB...")
        inserted = upsert_programs(valid_records)
        print(f"[OK] {inserted} programas insertados/actualizados.")
    else:
        print(f"\n[OK] Validacion exitosa. {len(valid_records)} programas listos para importar.")
        print("     Re-correr con --commit para insertar en DB.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
