"""GH-LOCAL-CLIENT-CATALOG · Import del xlsx del cliente (2026-05-28).

Parsea ``Resumen nuevo contratos - Google Sheets - Trabajar sobre este.xlsx``
y puebla la tabla ``institutions_catalog``.

Uso:
    python scripts/import_institutions.py <ruta-al-xlsx>              # dry-run
    python scripts/import_institutions.py <ruta-al-xlsx> --commit     # escribe DB

Reglas:
    - Sheet ``Instituciones`` (1551 rows, 34 cols) es la fuente de verdad
      principal: trae contrato, comisiones, fechas, contactos, programas.
    - Sheet ``Instituciones Resumen`` (1316 rows, 11 cols) se usa como fallback
      para instituciones no presentes en ``Instituciones``.
    - Sheet ``Cancelados - No renovados`` se SKIPEA siempre (relaciones muertas).
    - Filas con ``#REF!`` o ``#N/A`` en country / institución se SKIPEAN.
    - Country normalizada: Canadá→Canada, USA/Estados Unidos→USA, Reino Unido→UK, etc.
    - Dedupe case-insensitive por nombre (la primera fuente gana).
    - Upsert por ``lower(name)`` (no por id) para que re-runs sean idempotentes.

Output:
    - Reporte de filas leídas / válidas / inserted / updated / errors / warnings.
    - En --commit: escribe a DB usando ``app.db.database.SessionLocal``.
"""
from __future__ import annotations

import argparse
import re
import sys
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

# Ensure project root is importable when running ``python scripts/...`` from repo root.
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl no instalado. Ejecutar: pip install openpyxl")

warnings.filterwarnings("ignore")


# ---------------------------------------------------------------------------
# Normalization tables
# ---------------------------------------------------------------------------

# Map de country crudo → canónico. Acentos / variantes / inglés/español → un único valor.
COUNTRY_MAP = {
    "australia": "Australia",
    "usa": "USA",
    "estados unidos": "USA",
    "united states": "USA",
    "us": "USA",
    "canada": "Canada",
    "canadá": "Canada",
    "reino unido": "UK",
    "united kingdom": "UK",
    "uk": "UK",
    "england": "UK",
    "inglaterra": "UK",
    "nueva zelanda": "New Zealand",
    "new zealand": "New Zealand",
    "españa": "Spain",
    "spain": "Spain",
    "alemania": "Germany",
    "germany": "Germany",
    "francia": "France",
    "france": "France",
    "italia": "Italy",
    "italy": "Italy",
    "irlanda": "Ireland",
    "ireland": "Ireland",
    "malta": "Malta",
    "dubai": "UAE",
    "uae": "UAE",
    "internacional": "International",
    "international": "International",
}

# Filas que indican datos rotos (provenían de fórmulas que perdieron sus enlaces externos
# en el xlsx anterior). Cuando aparecen como country o name, descartamos la fila.
BROKEN_MARKERS = {"#ref!", "#n/a", "#name?", "#value!", "", "none"}

VALID_CATEGORIES = {
    "Universidad",
    "College Privado",
    "Instituto Idiomas",
    "College Público",
    "High School",
    "Proveedor",
    "Camps",
    "Polytechnic",
    "Summer School",
    "Business School",
}


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------


@dataclass
class ImportReport:
    sheet: str
    total_rows: int = 0
    skipped_empty: int = 0
    skipped_broken: int = 0
    skipped_dedupe: int = 0
    new_records: int = 0
    enriched_records: int = 0
    warnings: List[Tuple[int, str, str]] = field(default_factory=list)

    def add_warning(self, row: int, fld: str, msg: str) -> None:
        self.warnings.append((row, fld, msg))


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _looks_broken(v: Any) -> bool:
    s = _str(v)
    if s is None:
        return False  # empty handled separately
    return s.lower() in BROKEN_MARKERS or s.startswith("#")


def _norm_country(raw: Any) -> Optional[str]:
    s = _str(raw)
    if s is None or _looks_broken(s):
        return None
    return COUNTRY_MAP.get(s.lower(), s)  # passthrough if not in map


def _norm_category(raw: Any) -> Optional[str]:
    s = _str(raw)
    if s is None or _looks_broken(s):
        return None
    # Loose match: capitalize first letter of each significant word
    lower = s.lower()
    for cat in VALID_CATEGORIES:
        if cat.lower() == lower:
            return cat
    return s  # passthrough


def _parse_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _str(v)
    if s is None:
        return None
    # Try ISO first
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_programs(*vals: Any) -> List[str]:
    out: List[str] = []
    for v in vals:
        s = _str(v)
        if s is None or _looks_broken(s):
            continue
        out.append(s)
    return out


def _parse_commissions(c1: Any, d1: Any, c2: Any, d2: Any, c3: Any, d3: Any, c4: Any, d4: Any) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for value, desc in ((c1, d1), (c2, d2), (c3, d3), (c4, d4)):
        sv = _str(value)
        sd = _str(desc)
        if sv is None and sd is None:
            continue
        out.append({"value": sv, "description": sd})
    return out


def _normalize_dedupe_key(name: str) -> str:
    """Lowercase + collapse whitespace; used to dedupe across sheets."""
    return re.sub(r"\s+", " ", name.strip().lower())


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------


def parse_instituciones(ws, report: ImportReport) -> List[Dict[str, Any]]:
    """Parse the rich `Instituciones` sheet (34 cols).

    Column layout (0-indexed from openpyxl):
        0  marker ("Ok revisado ...")
        1  A NOMBRE DE QUIEN EL CONTRATO
        2  Institución
        3  Categoría
        4  Territorios
        5  Website
        6  Pertenece a grupo (Si/No)
        7  Nombre del Grupo
        8  Country
        9  Ciudad
        10 Programas que puedo vender bajo contrato
        11 Starting date
        12 End date
        13 Contact
        14 Email
        15 Programa 1
        16 Programa 2
        17 Programa 3
        18 Edvisor
        19 Estado
        20 Agreement Status
        21 Commission 1
        22 Commission 1 Description
        23 Commission 2
        24 Commission 2 Description
        25 Commission 3
        26 Commission 3 Description
        27 Commission 4
        28 Commission 4 Description
        29 Bonus
        30 Note
        31 Territory
    """
    out: List[Dict[str, Any]] = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    report.total_rows = len(rows)

    for i, row in enumerate(rows, start=2):
        if not row or all(v is None for v in row):
            report.skipped_empty += 1
            continue
        name = _str(row[2] if len(row) > 2 else None)
        if name is None:
            report.skipped_empty += 1
            continue
        if _looks_broken(name):
            report.skipped_broken += 1
            report.add_warning(i, "name", f"broken: {name!r}")
            continue

        country_raw = _str(row[8] if len(row) > 8 else None)
        country = _norm_country(country_raw)
        if country is None and country_raw is not None and _looks_broken(country_raw):
            report.add_warning(i, "country", f"broken: {country_raw!r}")

        record = {
            "name": name,
            "category": _norm_category(row[3] if len(row) > 3 else None),
            "country": country,
            "country_raw": country_raw,
            "city": _str(row[9] if len(row) > 9 else None),
            "partner_group": _str(row[7] if len(row) > 7 else None),
            "programs_offered": _parse_programs(
                row[15] if len(row) > 15 else None,
                row[16] if len(row) > 16 else None,
                row[17] if len(row) > 17 else None,
                row[10] if len(row) > 10 else None,
            ),
            "agreement_status": _str(row[20] if len(row) > 20 else None) or _str(row[19] if len(row) > 19 else None),
            "starting_date": _parse_date(row[11] if len(row) > 11 else None),
            "end_date": _parse_date(row[12] if len(row) > 12 else None),
            "contact_name": _str(row[13] if len(row) > 13 else None),
            "contact_email": _str(row[14] if len(row) > 14 else None),
            "website": _str(row[5] if len(row) > 5 else None),
            "territories": _str(row[4] if len(row) > 4 else None),
            "commissions": _parse_commissions(
                row[21] if len(row) > 21 else None,
                row[22] if len(row) > 22 else None,
                row[23] if len(row) > 23 else None,
                row[24] if len(row) > 24 else None,
                row[25] if len(row) > 25 else None,
                row[26] if len(row) > 26 else None,
                row[27] if len(row) > 27 else None,
                row[28] if len(row) > 28 else None,
            ),
            "source_sheet": "Instituciones",
            "active": True,
            "raw": {f"col_{ci}": (str(v) if not isinstance(v, (datetime, date)) else v.isoformat()) for ci, v in enumerate(row) if v is not None},
        }
        out.append(record)

    return out


def parse_resumen(ws, report: ImportReport) -> List[Dict[str, Any]]:
    """Parse `Instituciones Resumen` (11 cols). Used as fallback."""
    out: List[Dict[str, Any]] = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    report.total_rows = len(rows)
    for i, row in enumerate(rows, start=2):
        if not row or all(v is None for v in row):
            report.skipped_empty += 1
            continue
        name = _str(row[0] if len(row) > 0 else None)
        if name is None:
            report.skipped_empty += 1
            continue
        if _looks_broken(name):
            report.skipped_broken += 1
            continue
        country_raw = _str(row[4] if len(row) > 4 else None)
        record = {
            "name": name,
            "category": _norm_category(row[1] if len(row) > 1 else None),
            "country": _norm_country(country_raw),
            "country_raw": country_raw,
            "city": _str(row[5] if len(row) > 5 else None),
            "partner_group": _str(row[3] if len(row) > 3 else None),
            "programs_offered": _parse_programs(
                row[6] if len(row) > 6 else None,
                row[7] if len(row) > 7 else None,
                row[8] if len(row) > 8 else None,
                row[9] if len(row) > 9 else None,
                row[10] if len(row) > 10 else None,
            ),
            "agreement_status": None,
            "starting_date": None,
            "end_date": None,
            "contact_name": None,
            "contact_email": None,
            "website": None,
            "territories": None,
            "commissions": None,
            "source_sheet": "Instituciones Resumen",
            "active": True,
            "raw": {f"col_{ci}": (str(v) if not isinstance(v, (datetime, date)) else v.isoformat()) for ci, v in enumerate(row) if v is not None},
        }
        out.append(record)
    return out


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------


def dedupe_and_merge(records_by_source: List[Tuple[str, List[Dict[str, Any]]]]) -> Tuple[List[Dict[str, Any]], Dict[str, ImportReport]]:
    """First source wins; later sources only add NEW names (no overwrites)."""
    seen: Dict[str, Dict[str, Any]] = {}
    reports: Dict[str, ImportReport] = {}
    for sheet_name, recs in records_by_source:
        rep = reports.setdefault(sheet_name, ImportReport(sheet=sheet_name))
        for rec in recs:
            key = _normalize_dedupe_key(rec["name"])
            if key in seen:
                rep.skipped_dedupe += 1
                continue
            seen[key] = rec
            rep.new_records += 1
    return list(seen.values()), reports


def commit_to_db(records: List[Dict[str, Any]]) -> Tuple[int, int]:
    """Upsert records into ``institutions_catalog``. Match by lower(name).

    Returns (inserted, updated).
    """
    try:
        from sqlalchemy import func, inspect
        from app.db.database import SessionLocal, engine
        from app.db.models import InstitutionCatalog
    except Exception as exc:
        print(f"[ERROR] No se pudo importar SQLAlchemy/DB: {exc}", file=sys.stderr)
        return (0, 0)

    inspector = inspect(engine)
    if "institutions_catalog" not in inspector.get_table_names():
        print("[ERROR] Tabla institutions_catalog no existe. Correr alembic upgrade head primero.", file=sys.stderr)
        return (0, 0)

    inserted = 0
    updated = 0
    with SessionLocal() as session:
        # Build name → existing row map once (avoids N queries)
        existing = {
            _normalize_dedupe_key(row.name): row
            for row in session.query(InstitutionCatalog).all()
        }
        for rec in records:
            key = _normalize_dedupe_key(rec["name"])
            obj = existing.get(key)
            if obj is None:
                obj = InstitutionCatalog(**rec)
                session.add(obj)
                inserted += 1
            else:
                # Only enrich NULL fields; never overwrite existing values
                for k, v in rec.items():
                    if v is None:
                        continue
                    if getattr(obj, k, None) in (None, [], {}):
                        setattr(obj, k, v)
                updated += 1
        session.commit()
    return inserted, updated


def main() -> int:
    parser = argparse.ArgumentParser(description="Import institutions catalog from client xlsx")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--commit", action="store_true", help="Escribir a DB (default es dry-run)")
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    if not args.xlsx_path.exists():
        print(f"[ERROR] No existe: {args.xlsx_path}", file=sys.stderr)
        return 2

    print(f"[INFO] Cargando {args.xlsx_path}...")
    wb = load_workbook(args.xlsx_path, data_only=True)
    available_sheets = wb.sheetnames

    sources: List[Tuple[str, List[Dict[str, Any]]]] = []
    sheet_reports: Dict[str, ImportReport] = {}

    if "Instituciones" in available_sheets:
        rep = ImportReport(sheet="Instituciones")
        sheet_reports["Instituciones"] = rep
        recs = parse_instituciones(wb["Instituciones"], rep)
        sources.append(("Instituciones", recs))
        print(f"[INFO]   Instituciones: {len(recs)} válidas (total {rep.total_rows} · empty {rep.skipped_empty} · broken {rep.skipped_broken})")
    else:
        print("[WARN] No se encontró sheet 'Instituciones'.", file=sys.stderr)

    if "Instituciones Resumen" in available_sheets:
        rep = ImportReport(sheet="Instituciones Resumen")
        sheet_reports["Instituciones Resumen"] = rep
        recs = parse_resumen(wb["Instituciones Resumen"], rep)
        sources.append(("Instituciones Resumen", recs))
        print(f"[INFO]   Instituciones Resumen: {len(recs)} válidas (total {rep.total_rows} · empty {rep.skipped_empty} · broken {rep.skipped_broken})")

    if "Cancelados - No renovados" in available_sheets:
        cancelled_count = sum(
            1 for row in wb["Cancelados - No renovados"].iter_rows(min_row=2, values_only=True)
            if row and any(v is not None for v in row)
        )
        print(f"[INFO]   Cancelados - No renovados: SKIPPED ({cancelled_count} rows ignoradas por diseño)")

    merged, merge_reports = dedupe_and_merge(sources)
    for k, v in merge_reports.items():
        if k in sheet_reports:
            sheet_reports[k].skipped_dedupe = v.skipped_dedupe
            sheet_reports[k].new_records = v.new_records

    print("\n=== Reporte de import ===")
    print(f"Sheets procesados:    {len(sheet_reports)}")
    for name, rep in sheet_reports.items():
        print(f"  · {name}: new={rep.new_records} dedupe-skipped={rep.skipped_dedupe} broken={rep.skipped_broken} empty={rep.skipped_empty} warnings={len(rep.warnings)}")
    print(f"Total únicos a importar: {len(merged)}")

    # Country breakdown
    countries: Dict[str, int] = {}
    for rec in merged:
        countries[rec.get("country") or "?"] = countries.get(rec.get("country") or "?", 0) + 1
    print(f"\nTop countries (canónicos):")
    for k, v in sorted(countries.items(), key=lambda x: -x[1])[:15]:
        print(f"  {k}: {v}")

    if args.verbose:
        print("\nWarnings:")
        for rep in sheet_reports.values():
            for r, f, m in rep.warnings[:20]:
                print(f"  [{rep.sheet}] row {r} · {f}: {m}")

    if not args.commit:
        print("\n[OK] Dry-run · NO se escribió a DB. Re-correr con --commit para persistir.")
        return 0

    print(f"\n[INFO] Escribiendo {len(merged)} registros a institutions_catalog...")
    inserted, updated = commit_to_db(merged)
    print(f"[OK] inserted={inserted} updated={updated}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
